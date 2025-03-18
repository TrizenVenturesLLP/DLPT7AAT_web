from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import face_recognition
import cv2
import numpy as np
import os
import base64
from datetime import date
from openpyxl import Workbook, load_workbook
from gaze_tracking import GazeTracking

app = Flask(__name__)
CORS(app)

# Initialize gaze tracking
gaze = GazeTracking()

# Initialize arrays for known face encodings and names
known_face_encodings = []
known_face_names = []

# Load face encodings from data folder
data_folder = "data"
for file in os.listdir(data_folder):
    if file.lower().endswith(('png', 'jpg', 'jpeg')):
        file_path = os.path.join(data_folder, file)
        try:
            image = face_recognition.load_image_file(file_path)
            face_encodings = face_recognition.face_encodings(image)
            if face_encodings:
                known_face_encodings.append(face_encodings[0])
                known_face_names.append(os.path.splitext(file)[0])
        except Exception as e:
            print(f"Error processing file {file}: {e}")

@app.route('/api/process-frame', methods=['POST'])
def process_frame():
    try:
        # Get the base64 image from the request
        data = request.json
        base64_image = data['frame'].split(',')[1]
        image_data = base64.b64decode(base64_image)
        
        # Convert to numpy array
        nparr = np.frombuffer(image_data, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        # Convert BGR to RGB
        rgb_frame = frame[:, :, ::-1]
        
        # Process gaze tracking
        gaze.refresh(frame)
        engagement_score = 100  # Default engagement score
        engagement_remarks = "Actively participating"
        
        # Determine engagement based on gaze
        if gaze.is_blinking():
            engagement_score -= 30
            engagement_remarks = "Student appears to be sleeping"
        elif gaze.is_right() or gaze.is_left():
            engagement_score -= 20
            engagement_remarks = "Student is distracted"
        
        # Find faces in the frame
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        
        face_names = []
        
        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"
            
            if True in matches:
                best_match_index = np.argmin(face_recognition.face_distance(known_face_encodings, face_encoding))
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]
                    
            face_names.append(name)
            
            # Update attendance in Excel
            if name != "Unknown":
                update_attendance(name, engagement_score, engagement_remarks)
        
        return jsonify({
            'faces': face_names,
            'engagement': engagement_score,
            'remarks': engagement_remarks,
            'gaze_status': gaze.get_gaze_status()
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def update_attendance(name, engagement, remarks):
    try:
        today = date.today().strftime("%Y-%m-%d")
        filename = "attendance.xlsx"
        
        if not os.path.exists(filename):
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Attendance"
            sheet.append(["Date", "Name", "Status", "Engagement", "Remarks"])
            wb.save(filename)

        wb = load_workbook(filename)
        sheet = wb.active
        sheet.append([today, name, "Present", engagement, remarks])  # Append new data
        wb.save(filename)

    except Exception as e:
        print(f"Error updating attendance: {e}")

@app.route('/api/download-attendance', methods=['GET'])
def download_attendance():
    try:
        return send_file(
            'attendance.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'attendance_{date.today().strftime("%Y-%m-%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/get-attendance', methods=['GET'])
def get_attendance():
    try:
        if not os.path.exists('attendance.xlsx'):
            return jsonify([])

        wb = load_workbook('attendance.xlsx')
        sheet = wb.active

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            data.append({
                'date': row[0],
                'name': row[1],
                'status': row[2],
                'engagement': row[3],
                'remarks': row[4]
            })

        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
